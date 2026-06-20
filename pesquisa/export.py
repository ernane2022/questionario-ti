import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _escrever_aba(ws, colunas, queryset):
    fill = PatternFill("solid", fgColor="1E4D8C")
    font = Font(color="FFFFFF", bold=True)
    for col_idx, (label, _) in enumerate(colunas, 1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")
    for row_idx, obj in enumerate(queryset, 2):
        for col_idx, (_, attr) in enumerate(colunas, 1):
            val = getattr(obj, attr)
            if callable(val):
                val = val()
            if hasattr(val, "strftime"):
                val = val.strftime("%d/%m/%Y %H:%M")
            if isinstance(val, bool):
                val = "Sim" if val else "Não"
            ws.cell(row=row_idx, column=col_idx, value=str(val) if val else "")
    for col_idx in range(1, len(colunas) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25
    ws.freeze_panes = "A2"


def exportar_excel_multi(prof_qs, est_qs, pub_qs):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Profissionais"
    _escrever_aba(ws1, [
        ("ID", "id"), ("Data", "criado_em"),
        ("Faixa Etária", "get_faixa_etaria_display"),
        ("Gênero", "get_genero_display"),
        ("Escolaridade", "get_escolaridade_display"),
        ("Estado", "estado"), ("Cidade", "cidade"),
        ("Mercado", "get_percebe_mercado_display"),
        ("Área", "get_area_atuacao_display"),
        ("Tempo", "get_tempo_atuacao_display"),
        ("Regime", "get_regime_trabalho_display"),
        ("Remoto", "trabalha_remoto"),
        ("Atualização", "atualizacao_constante"),
        ("Stack", "stack_principal"),
        ("Comp. Técnicas", "competencias_tecnicas"),
        ("Comp. Comportamentais", "competencias_comportamentais"),
        ("Dificuldade", "maior_dificuldade"),
        ("Conselho", "conselho_iniciante"),
        ("Percepção", "percepcao_profissao"),
    ], prof_qs)

    ws2 = wb.create_sheet("Estudantes")
    _escrever_aba(ws2, [
        ("ID", "id"), ("Data", "criado_em"),
        ("Faixa Etária", "get_faixa_etaria_display"),
        ("Gênero", "get_genero_display"),
        ("Escolaridade", "get_escolaridade_display"),
        ("Estado", "estado"), ("Cidade", "cidade"),
        ("Mercado", "get_percebe_mercado_display"),
        ("Curso", "curso"), ("Instituição", "instituicao"),
        ("Semestre", "get_semestre_display"),
        ("Turno", "get_turno_display"),
        ("Motivo", "get_motivo_escolha_display"),
        ("Estágio", "possui_estagio"),
        ("Comp. Técnicas", "competencias_tecnicas"),
        ("Comp. Comportamentais", "competencias_comportamentais"),
        ("Dificuldade", "dificuldade_academica"),
        ("Expectativa", "expectativa_carreira"),
        ("Conselho", "conselho_iniciante"),
    ], est_qs)

    ws3 = wb.create_sheet("Público Geral")
    _escrever_aba(ws3, [
        ("ID", "id"), ("Data", "criado_em"),
        ("Faixa Etária", "get_faixa_etaria_display"),
        ("Gênero", "get_genero_display"),
        ("Escolaridade", "get_escolaridade_display"),
        ("Estado", "estado"), ("Cidade", "cidade"),
        ("Mercado", "get_percebe_mercado_display"),
        ("Freq. Uso", "get_frequencia_uso_tecnologia_display"),
        ("Confiança", "get_confianca_sistemas_display"),
        ("Associa TI", "get_associa_ti_a_display"),
        ("Conhece Prof. TI", "conhece_profissional_ti"),
        ("Impacto IA", "impacto_ia"),
        ("Percepção", "percepcao_profissao"),
        ("Uso Cotidiano", "uso_tecnologia_cotidiano"),
        ("Preocupação", "preocupacao_digital"),
    ], pub_qs)

    return wb